"""Generate company profile .xlsx files using the Claude API.

For each company in list_of_companies_to_seed.txt:
  Step 1 — Research the company   (raw_profile_data_prompt.txt)
  Step 2 — Format into schema     (formatting_data_prompt.txt)
  Step 3 — Build .xlsx locally    (generate_xlsx_profile.txt spec)
  Saves to company_profiles/<COMPANY_NAME>.xlsx

Run from the project root:
    python -m company_culture_LLM.generate_profiles
"""

from __future__ import annotations

import os
import re
import sys
import time
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

load_dotenv()

ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
COMPANY_PROFILES = ROOT / "company_profiles"
COMPANY_LIST = ROOT / "company_culture_LLM" / "list_of_companies_to_seed.txt"

MODEL = "claude-opus-4-6"

# ─────────────────────────────────────────────
# XLSX style constants
# ─────────────────────────────────────────────
DARK_NAVY = "1F3864"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
LABEL_GRAY = "595959"


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


# ─────────────────────────────────────────────
# File helpers
# ─────────────────────────────────────────────

def read_prompt(filename: str) -> str:
    return (DOCS / filename).read_text(encoding="utf-8")


def load_companies() -> list[str]:
    text = COMPANY_LIST.read_text(encoding="utf-8")
    return [line.strip() for line in text.splitlines() if line.strip()]


def xlsx_safe_name(company: str) -> str:
    """Convert company name to the uppercase underscore filename used in company_profiles/."""
    safe = re.sub(r"[^\w\s\(\)]", "", company).strip()
    safe = re.sub(r"\s+", "_", safe).upper()
    return safe


# ─────────────────────────────────────────────
# Claude API calls
# ─────────────────────────────────────────────

def step1_research(client: anthropic.Anthropic, company: str, prompt_template: str) -> str:
    """Call Claude with the raw research prompt for a specific company."""
    prompt = prompt_template.replace("[COMPANY NAME]", company)
    msg = client.messages.create(
        model=MODEL,
        max_tokens=8000,
        messages=[{"role": "user", "content": prompt}],
    )
    return msg.content[0].text


def step2_format(client: anthropic.Anthropic, research: str, format_prompt: str) -> str:
    """Pass the research as prior context, then ask Claude to format it."""
    # We present the research as the assistant's prior turn so the formatter
    # can reference "what was established in Step 1" exactly as the prompt expects.
    messages = [
        {
            "role": "user",
            "content": (
                "Please research this company and provide a complete intelligence report "
                "following the sections and structure in your instructions."
            ),
        },
        {
            "role": "assistant",
            "content": research,
        },
        {
            "role": "user",
            "content": format_prompt,
        },
    ]
    msg = client.messages.create(
        model=MODEL,
        max_tokens=4000,
        messages=messages,
    )
    return msg.content[0].text


# ─────────────────────────────────────────────
# Step 2 output parser
# ─────────────────────────────────────────────

# Matches lines that start a new field, e.g.:
#   overview.name: JPMorgan Chase & Co.
#   executives[0].name: Jamie Dimon
_FIELD_RE = re.compile(
    r"^([a-zA-Z_][a-zA-Z_0-9]*(?:\[\d+\])?(?:\.[a-zA-Z_][a-zA-Z_0-9]*(?:\[\d+\])?)+)\s*:\s*(.*)"
)


def parse_structured_output(text: str) -> dict[str, str]:
    """Parse the flat key: value output from Step 2 into a dict.

    Handles multi-line values — continuation lines are appended until
    the next field key is detected.
    """
    result: dict[str, str] = {}
    current_key: str | None = None
    current_lines: list[str] = []

    for line in text.splitlines():
        m = _FIELD_RE.match(line)
        if m:
            # Flush previous key
            if current_key is not None:
                result[current_key] = "\n".join(current_lines).strip()
            current_key = m.group(1)
            current_lines = [m.group(2)]
        else:
            if current_key is not None:
                current_lines.append(line)

    # Flush last key
    if current_key is not None:
        result[current_key] = "\n".join(current_lines).strip()

    return result


# ─────────────────────────────────────────────
# XLSX builder  (implements generate_xlsx_profile.txt spec)
# ─────────────────────────────────────────────

def _add_title(ws, text: str) -> None:
    row = ws.max_row + 1 if ws.max_row else 1
    ws.row_dimensions[row].height = 24
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    cell.fill = _fill(DARK_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_section(ws, text: str) -> None:
    row = ws.max_row + 1
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = _fill(DARK_NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _add_field(ws, label: str, value: str) -> None:
    row = ws.max_row + 1
    ws.row_dimensions[row].height = 15

    a = ws.cell(row=row, column=1, value=label)
    a.font = Font(name="Arial", size=9, bold=True, color=LABEL_GRAY)
    a.fill = _fill(LIGHT_GRAY)
    a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    a.border = _thin_border()

    b = ws.cell(row=row, column=2, value=value)
    b.font = Font(name="Arial", size=9, color="000000")
    b.fill = _fill(WHITE)
    b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    b.border = _thin_border()


def _resolve(data: dict, key: str) -> str:
    """Return the display string for a field, applying boolean conversions."""
    raw = data.get(key, "")

    # Boolean fields
    if key == "interview.system_design":
        low = str(raw).strip().lower()
        if low == "true":
            return "Yes"
        if low == "false":
            return "No"
        return ""

    if key == "internship.program_exists":
        low = str(raw).strip().lower()
        if low == "true":
            return "Yes"
        if low == "false":
            return "No"
        return ""

    # None / "None" → empty string
    if raw is None or str(raw).strip().lower() == "none":
        return ""

    return str(raw).strip()


def build_xlsx(data: dict, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Profile"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    ws.freeze_panes = "A2"

    def f(key: str) -> str:
        return _resolve(data, key)

    _add_title(ws, "COMPANY PROFILE")

    # ── COMPANY OVERVIEW ──────────────────────
    _add_section(ws, "COMPANY OVERVIEW")
    _add_field(ws, "Name",              f("overview.name"))
    _add_field(ws, "Founded",           f("overview.founded"))
    _add_field(ws, "Website",           f("overview.website"))
    _add_field(ws, "Industry",          f("overview.industry"))
    _add_field(ws, "Size",              f("overview.size"))
    _add_field(ws, "Headquarters",      f("overview.headquarters"))
    _add_field(ws, "Ownership",         f("overview.ownership"))
    _add_field(ws, "Revenue (est.)",    f("overview.revenue_est"))
    _add_field(ws, "Headcount (est.)",  f("overview.headcount_est"))

    # ── EXECUTIVE LEADERSHIP ──────────────────
    _add_section(ws, "EXECUTIVE LEADERSHIP")
    for i in range(5):
        p = f"executives[{i}]"
        _add_field(ws, f"Executive {i+1} — Name",        f(f"{p}.name"))
        _add_field(ws, f"Executive {i+1} — Title",       f(f"{p}.title"))
        _add_field(ws, f"Executive {i+1} — LinkedIn URL", f(f"{p}.linkedin_url"))

    # ── CULTURE & VALUES ──────────────────────
    _add_section(ws, "CULTURE & VALUES")
    _add_field(ws, "Culture Narrative (Company Says)",   f("culture.culture_narrative_company"))
    _add_field(ws, "Culture Narrative (Employees Say)",  f("culture.culture_narrative_employees"))
    _add_field(ws, "Core Values",                        f("culture.core_values"))
    _add_field(ws, "Work Environment",                   f("culture.work_environment"))
    _add_field(ws, "Pace",                               f("culture.pace"))
    _add_field(ws, "Empathy (0.0–1.0)",                  f("culture.empathy"))
    _add_field(ws, "Creative Drive (0.0–1.0)",           f("culture.creative_drive"))
    _add_field(ws, "Adaptability (0.0–1.0)",             f("culture.adaptability"))
    _add_field(ws, "Futuristic (0.0–1.0)",               f("culture.futuristic"))
    _add_field(ws, "Harmony (0.0–1.0)",                  f("culture.harmony"))
    _add_field(ws, "Data Orientation (0.0–1.0)",         f("culture.data_orientation"))

    # ── COMPENSATION ──────────────────────────
    _add_section(ws, "COMPENSATION")
    _add_field(ws, "New Grad SWE Total", f("compensation.new_grad_swe_total"))
    _add_field(ws, "Mid-Level Total",    f("compensation.mid_level_total"))
    _add_field(ws, "Senior Total",       f("compensation.senior_total"))
    _add_field(ws, "Base Range",         f("compensation.base_range"))
    _add_field(ws, "Equity Notes",       f("compensation.equity_notes"))
    _add_field(ws, "Bonus",              f("compensation.bonus"))
    _add_field(ws, "Levels",             f("compensation.levels"))
    _add_field(ws, "Notes",              f("compensation.notes"))

    # ── INTERVIEW PROCESS ─────────────────────
    _add_section(ws, "INTERVIEW PROCESS")
    _add_field(ws, "Difficulty",           f("interview.difficulty"))
    _add_field(ws, "Process Overview",     f("interview.process_overview"))
    _add_field(ws, "Coding Focus",         f("interview.coding_focus"))
    _add_field(ws, "System Design (Yes/No)", f("interview.system_design"))
    _add_field(ws, "Behavioral",           f("interview.behavioral"))
    _add_field(ws, "Avg Rounds",           f("interview.avg_rounds"))
    _add_field(ws, "Timeline",             f("interview.timeline"))
    _add_field(ws, "Notes",                f("interview.notes"))

    # ── REMOTE / HYBRID POLICY ────────────────
    _add_section(ws, "REMOTE / HYBRID POLICY")
    _add_field(ws, "Policy",         f("remote_policy.policy"))
    _add_field(ws, "Remote Friendly", f("remote_policy.remote_friendly"))
    _add_field(ws, "Details",        f("remote_policy.details"))

    # ── GLASSDOOR / EMPLOYEE RATINGS ──────────
    _add_section(ws, "GLASSDOOR / EMPLOYEE RATINGS")
    _add_field(ws, "Overall Rating (/ 5.0)",  f("glassdoor.overall_rating"))
    _add_field(ws, "CEO Approval",            f("glassdoor.ceo_approval"))
    _add_field(ws, "Recommend to Friend",     f("glassdoor.recommend_to_friend"))
    _add_field(ws, "Work-Life Balance",       f("glassdoor.work_life_balance"))
    _add_field(ws, "Compensation Score",      f("glassdoor.compensation_score"))
    _add_field(ws, "Culture Score",           f("glassdoor.culture_score"))
    _add_field(ws, "Key Positives",           f("glassdoor.key_positives"))
    _add_field(ws, "Key Negatives",           f("glassdoor.key_negatives"))
    _add_field(ws, "Source",                  f("glassdoor.source"))

    # ── RECENT NEWS & CONTROVERSIES ───────────
    _add_section(ws, "RECENT NEWS & CONTROVERSIES")
    _add_field(ws, "Controversy Level", f("news.controversy_level"))
    _add_field(ws, "Key Topics",        f("news.key_topics"))
    _add_field(ws, "Recent News",       f("news.recent_news"))

    # ── RECRUITING & CONTACTS ─────────────────
    _add_section(ws, "RECRUITING & CONTACTS")
    _add_field(ws, "Key Programs",             f("recruiting.key_programs"))
    _add_field(ws, "University Recruiting URL", f("recruiting.university_recruiting_url"))
    _add_field(ws, "General Jobs URL",          f("recruiting.general_jobs_url"))
    _add_field(ws, "LinkedIn Search Tip",       f("recruiting.linkedin_search_tip"))
    _add_field(ws, "Recruiter Tip",             f("recruiting.recruiter_tip"))

    # ── TECH STACK ────────────────────────────
    _add_section(ws, "TECH STACK")
    _add_field(ws, "Frontend",          f("tech_stack.frontend"))
    _add_field(ws, "Backend",           f("tech_stack.backend"))
    _add_field(ws, "Infrastructure",    f("tech_stack.infrastructure"))
    _add_field(ws, "Data / ML",         f("tech_stack.data_ml"))
    _add_field(ws, "DevTools / CI/CD",  f("tech_stack.devtools_cicd"))
    _add_field(ws, "Notable OSS",       f("tech_stack.notable_oss"))
    _add_field(ws, "Primary Languages", f("tech_stack.primary_languages"))
    _add_field(ws, "Notes",             f("tech_stack.notes"))

    # ── INTERNSHIP PROGRAM ────────────────────
    _add_section(ws, "INTERNSHIP PROGRAM")
    _add_field(ws, "Program Exists (Yes/No)", f("internship.program_exists"))
    _add_field(ws, "Program Name",            f("internship.program_name"))
    _add_field(ws, "Duration (weeks)",        f("internship.duration_weeks"))
    _add_field(ws, "Roles Available",         f("internship.roles_available"))
    _add_field(ws, "Weekly Comp (USD)",       f("internship.weekly_comp_usd"))
    _add_field(ws, "Return Offer Rate",       f("internship.return_offer_rate"))
    _add_field(ws, "Pipeline to Full-Time",   f("internship.pipeline_to_fulltime"))
    _add_field(ws, "Application Timeline",    f("internship.application_timeline"))
    _add_field(ws, "Intern Headcount Est.",   f("internship.intern_headcount_est"))
    _add_field(ws, "Notes",                   f("internship.notes"))

    # ── WORKFORCE SIGNALS ─────────────────────
    _add_section(ws, "WORKFORCE SIGNALS")
    _add_field(ws, "WARN Activity",          f("workforce_signals.warn_activity"))
    _add_field(ws, "WARN Risk Level",        f("workforce_signals.warn_risk_level"))
    _add_field(ws, "WARN Analyst Note",      f("workforce_signals.warn_analyst_note"))
    _add_field(ws, "Reddit / Glassdoor Themes", f("workforce_signals.reddit_glassdoor_themes"))
    _add_field(ws, "LinkedIn Signals",       f("workforce_signals.linkedin_signals"))
    _add_field(ws, "News / Press Signals",   f("workforce_signals.news_press_signals"))
    _add_field(ws, "Overall Workforce Risk", f("workforce_signals.overall_workforce_risk"))
    _add_field(ws, "Signal Confidence",      f("workforce_signals.signal_confidence"))
    _add_field(ws, "Analyst Summary",        f("workforce_signals.analyst_summary"))

    COMPANY_PROFILES.mkdir(exist_ok=True)
    wb.save(output_path)


# ─────────────────────────────────────────────
# Per-company pipeline
# ─────────────────────────────────────────────

def process_company(
    client: anthropic.Anthropic,
    company: str,
    raw_prompt: str,
    format_prompt: str,
) -> None:
    safe = xlsx_safe_name(company)
    output_path = COMPANY_PROFILES / f"{safe}.xlsx"

    if output_path.exists():
        print(f"  [skip] {output_path.name} already exists — delete it to regenerate")
        return

    print(f"  [1/3] Researching...")
    research = step1_research(client, company, raw_prompt)

    print(f"  [2/3] Formatting into schema...")
    formatted = step2_format(client, research, format_prompt)

    print(f"  [3/3] Building xlsx...")
    data = parse_structured_output(formatted)
    build_xlsx(data, output_path)
    print(f"  ✓  Saved → {output_path.name}")


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

def main() -> None:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not set. Add it to .env or export it.")
        sys.exit(1)

    companies = load_companies()
    if not companies:
        print("No companies found in list_of_companies_to_seed.txt — add one per line.")
        sys.exit(0)

    print(f"Found {len(companies)} companies to process.\n")

    client = anthropic.Anthropic(api_key=api_key)
    raw_prompt = read_prompt("raw_profile_data_prompt.txt")
    format_prompt = read_prompt("formatting_data_prompt.txt")

    for idx, company in enumerate(companies, 1):
        print(f"\n{'═' * 60}")
        print(f"  {idx}/{len(companies)}  {company}")
        print(f"{'═' * 60}")
        try:
            process_company(client, company, raw_prompt, format_prompt)
        except anthropic.RateLimitError:
            print("  Rate limit hit — waiting 60 s before retrying...")
            time.sleep(60)
            process_company(client, company, raw_prompt, format_prompt)
        except Exception as exc:
            print(f"  ERROR: {exc}")
            print("  Skipping this company and continuing...")

        # Brief pause between companies to avoid hammering the API
        if idx < len(companies):
            time.sleep(3)

    print(f"\nDone — processed {len(companies)} companies.")
    print(f"Files saved to: {COMPANY_PROFILES}")


if __name__ == "__main__":
    main()
