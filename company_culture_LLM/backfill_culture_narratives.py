"""backfill_culture_narratives.py

Reads docs/Original profiles with only class data.xlsx, calls Claude API to generate
culture_narrative_company and culture_narrative_employees for each company, then writes
a new spreadsheet for review.  Does NOT touch the database.

Usage:
    # Set key in environment
    ANTHROPIC_API_KEY=sk-ant-... python company_culture_LLM/backfill_culture_narratives.py

    # Or pass it as an argument
    python company_culture_LLM/backfill_culture_narratives.py --api-key sk-ant-...

Output:
    docs/Original profiles with culture narratives REVIEW.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

import anthropic
import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

INPUT_XLSX  = ROOT / "docs" / "Original profiles with only class data.xlsx"
OUTPUT_XLSX = ROOT / "docs" / "Original profiles with culture narratives REVIEW.xlsx"

# ---------------------------------------------------------------------------
# Section headers (same set as parse_xlsx.py)
# ---------------------------------------------------------------------------

SECTION_HEADERS = {
    "COMPANY OVERVIEW",
    "EXECUTIVE LEADERSHIP",
    "CULTURE & VALUES",
    "COMPENSATION",
    "INTERVIEW PROCESS",
    "REMOTE / HYBRID POLICY",
    "GLASSDOOR / EMPLOYEE RATINGS",
    "RECENT NEWS & CONTROVERSIES",
    "RECRUITING & CONTACTS",
    "TECH STACK",
    "INTERNSHIP PROGRAM",
    "WORKFORCE SIGNALS",
}


def _is_header_row(row: tuple) -> bool:
    val = row[0]
    return (
        isinstance(val, str)
        and val.startswith("  ")
        and all(v is None for v in row[1:])
    )


def _classify_header(val: str) -> tuple[str, str]:
    stripped = val.strip().upper()
    if stripped in SECTION_HEADERS:
        return "section", stripped
    return "company", val.strip()


# ---------------------------------------------------------------------------
# Parse spreadsheet into per-company context dicts
# ---------------------------------------------------------------------------

def _parse_companies(ws) -> list[dict]:
    """Return list of dicts, each with 'name', 'rows' (all raw rows), and 'culture_kv'."""
    companies: list[dict] = []
    current_company: str | None = None
    current_section: str | None = None
    current_rows: list[tuple] = []
    culture_kv: dict[str, str] = {}
    glassdoor_kv: dict[str, str] = {}
    workforce_kv: dict[str, str] = {}

    def _flush():
        if current_company:
            companies.append({
                "name": current_company,
                "rows": list(current_rows),
                "culture_kv": dict(culture_kv),
                "glassdoor_kv": dict(glassdoor_kv),
                "workforce_kv": dict(workforce_kv),
            })

    for row in ws.iter_rows(values_only=True):
        if _is_header_row(row):
            kind, name = _classify_header(row[0])
            if kind == "company":
                _flush()
                current_company = name
                current_section = None
                current_rows = [row]
                culture_kv = {}
                glassdoor_kv = {}
                workforce_kv = {}
            else:
                current_section = name
                current_rows.append(row)
        else:
            if current_company is not None:
                current_rows.append(row)
                # Accumulate key-value pairs for context sections
                label = row[0]
                value = row[1] if len(row) > 1 else None
                if label and value is not None:
                    kv_entry = (str(label).strip(), str(value).strip())
                    if current_section == "CULTURE & VALUES":
                        culture_kv[kv_entry[0]] = kv_entry[1]
                    elif current_section == "GLASSDOOR / EMPLOYEE RATINGS":
                        glassdoor_kv[kv_entry[0]] = kv_entry[1]
                    elif current_section == "WORKFORCE SIGNALS":
                        workforce_kv[kv_entry[0]] = kv_entry[1]

    _flush()
    return companies


# ---------------------------------------------------------------------------
# Build Claude prompt context for a company
# ---------------------------------------------------------------------------

def _build_context(company: dict) -> str:
    c = company["culture_kv"]
    g = company["glassdoor_kv"]
    w = company["workforce_kv"]

    lines = [f"Company: {company['name']}"]

    if c.get("Culture Narrative"):
        lines.append(f"Existing Culture Narrative: {c['Culture Narrative']}")
    if c.get("Core Values"):
        lines.append(f"Core Values: {c['Core Values']}")
    if c.get("Work Environment"):
        lines.append(f"Work Environment: {c['Work Environment']}")
    if c.get("Pace"):
        lines.append(f"Pace: {c['Pace']}")

    if g:
        lines.append("\nGlassdoor Ratings:")
        for k, v in g.items():
            lines.append(f"  {k}: {v}")

    if w:
        lines.append("\nWorkforce Signals:")
        for k, v in w.items():
            lines.append(f"  {k}: {v}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Call Claude API
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """\
You are a professional company culture analyst. Given company profile data, generate two distinct culture narratives.

Return ONLY valid JSON with exactly these two keys:
{
  "culture_narrative_company": "<2-4 sentences of official company framing from careers pages, CEO statements, employer branding. No employee sentiment.>",
  "culture_narrative_employees": "<4-7 sentences synthesized from Glassdoor, Indeed, Blind, TheLayoff.com, Wall Street Oasis, Reddit. Separate positives from negatives. Include paraphrased recurring themes or quotes. Note if sentiment varies by division, role, or manager quality.>"
}

Do not include any text outside the JSON object."""


def _call_claude(client: anthropic.Anthropic, company: dict) -> dict[str, str]:
    context = _build_context(company)

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": f"{context}\n\nGenerate the two culture narratives for this company."
            }
        ],
        system=SYSTEM_PROMPT,
    )

    raw = message.content[0].text.strip()

    # Strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        result = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"Claude returned invalid JSON for {company['name']}: {e}\n\nRaw:\n{raw}")

    if "culture_narrative_company" not in result or "culture_narrative_employees" not in result:
        raise ValueError(f"Missing expected keys in Claude response for {company['name']}: {result}")

    return result


# ---------------------------------------------------------------------------
# Write updated spreadsheet
# ---------------------------------------------------------------------------

def _write_output(companies: list[dict], narratives: dict[str, dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Profiles"

    # Column widths for readability
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 120

    bold = Font(bold=True)

    for company in companies:
        name = company["name"]
        generated = narratives.get(name, {})
        in_culture_section = False

        for row in company["rows"]:
            label = row[0]

            # Update section tracking BEFORE deciding what to write
            if _is_header_row(row):
                kind, sec_name = _classify_header(label)
                if kind == "section":
                    in_culture_section = (sec_name == "CULTURE & VALUES")
                elif kind == "company":
                    in_culture_section = False

            # Replace the single "Culture Narrative" row with the two new rows
            if (
                isinstance(label, str)
                and label.strip() == "Culture Narrative"
                and in_culture_section
            ):
                ws.append(["Culture Narrative (Company Says)", generated.get("culture_narrative_company", "")])
                ws.append(["Culture Narrative (Employees Say)", generated.get("culture_narrative_employees", "")])
                continue

            ws.append(list(row))

    # Re-apply bold after all rows written
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("  "):
            cell.font = bold

    wb.save(str(OUTPUT_XLSX))
    print(f"\nWrote: {OUTPUT_XLSX}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Backfill culture narratives via Claude API")
    parser.add_argument("--api-key", help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, skip API calls")
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("ANTHROPIC_API_KEY")
    if not api_key and not args.dry_run:
        sys.exit("Error: provide --api-key or set ANTHROPIC_API_KEY environment variable")

    if not INPUT_XLSX.exists():
        sys.exit(f"Error: input file not found: {INPUT_XLSX}")

    print(f"Reading: {INPUT_XLSX}")
    wb = openpyxl.load_workbook(str(INPUT_XLSX), read_only=True)
    ws = wb.active
    companies = _parse_companies(ws)
    print(f"Found {len(companies)} companies: {[c['name'] for c in companies]}")

    if args.dry_run:
        print("Dry run — skipping API calls.")
        return

    client = anthropic.Anthropic(api_key=api_key)
    narratives: dict[str, dict[str, str]] = {}

    for i, company in enumerate(companies, 1):
        name = company["name"]
        print(f"[{i}/{len(companies)}] Generating narratives for {name}...", end=" ", flush=True)
        try:
            result = _call_claude(client, company)
            narratives[name] = result
            print("done")
        except Exception as e:
            print(f"ERROR: {e}")
            narratives[name] = {
                "culture_narrative_company": f"[ERROR: {e}]",
                "culture_narrative_employees": f"[ERROR: {e}]",
            }

        # Small delay between API calls to avoid rate limits
        if i < len(companies):
            time.sleep(0.5)

    _write_output(companies, narratives)
    print("\nDone. Review the spreadsheet before seeding to the database.")


if __name__ == "__main__":
    main()
